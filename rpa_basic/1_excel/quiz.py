'''
Quiz) 여러분은 나도대학의 컴퓨터과 교수님입니다.
여러분이 가르치는 과목의 점수 비중은 다음과 같습니다.

- 출석 10
- 퀴즈1 10
- 퀴즈2 10
- 중간고사 20
- 기말고사 30
- 프로젝트 20

-----------------
총합 100

학생수 : 20

마지막 수업을 모두 마치고 이번 학기 학생들의 최종 성적을 검토하는 과정에서
퀴즈2 문제에 오류를 발견하여 모두 만점 처리를 하기로 하였습니다,
현재까지 작성된 최종 성적 데이터를 기준으로 아래와 같이 수정하시오.

1. 퀴즈 2 점수를 10으로 수정
2. H열에 총점 (SUM이용), I열에 성적 정보 추가
- 총점 90이상 A, 80이상 B, 70 이상 C, 나머지 D
3. 출석이 5미만인 학생은 총점 상관없이 F

*  최종 파일명 : scores.xlsx
'''

'''
문제 봤을때 가장 먼저 해야할 것
1. 컬럼 별로 범위에 맞게 랜덤으로 값 집어 넣기
2. 퀴즈 2는 10으로 고정 시키기
3. H열 수식 작성하기
4. I열 등급 매기는데...출석 조건주기
5. 마지막 파일명 저장
'''

from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl import load_workbook
from random import *

wb = Workbook()

ws = wb.active

#컬럼 이름 다 집어 넣기
#           1           2           3           4               5                   6               7           8    9      10                
ws.append(["학번", "출석 (10)", "퀴즈1 (10)", "퀴즈2 (10)", "중간고사 (20)", "기말고사 (30)", "프로젝트 (20)", "", "총점","등급"])

for alp in ["B", "C", "D", "E", "F", "G"]:
    ws.column_dimensions[alp].width = 15

# 학생 수에 따른 학번 입력하기
for row in range(2, 22):
    ws.cell(row=row, column=1, value=row-1)

# 출석 (10)점수 랜덤하게 집어넣기
for row in range(2,22):
    ws.cell(row=row, column=2, value=randint(0,10))

# 퀴즈1 (10)점수 랜덤하게 집어넣기
for row in range(2,22):
    ws.cell(row=row, column=3, value=randint(0,10))
    
# 퀴즈2 (10) 점수 랜덤하게 집어넣기 -> 오류로 인해 무조건 만점(10)주기
for row in range(2,22):
    ws.cell(row=row, column=4, value=10)

# 중간고사 (20) 점수 랜덤하게 집어넣기
for row in range(2,22):
    ws.cell(row=row, column=5, value=randint(0,20))

# 기말고사 (30) 점수 랜덤하게 집어넣기
for row in range(2,22):
    ws.cell(row=row, column=6, value=randint(0,30))

# 프로젝트 (20) 점수 랜덤하게 집어넣기
for row in range(2,22):
    ws.cell(row=row, column=7, value=randint(0,20))
    
#총점 계산 및 집어넣기
#ws["I2"] = "=SUM(B2:G2)"
for index in range(2,22):
    index = str(index)
    I = 'I'+index
    B = 'B'+index
    G = 'G'+index
    ws[I] = "=SUM("+B+":"+G+")"

wb.save("scores1.xlsx")

wb = load_workbook("scores1.xlsx")

ws = wb.active

#제가 개 벌레입니다. A~G
add_datas = 0
for row in range(2, 22):
    for col in range(2,8):
        add_datas += ws.cell(row=row, column=col).value
    #ws.cell(row=row, column=9, value= add_datas)
    #print(add_datas)
    if add_datas >= 90:
        ws.cell(row=row, column=10, value="A")
    elif add_datas >= 80:
        ws.cell(row=row, column=10, value="B")
    elif add_datas >= 70:
        ws.cell(row=row, column=10, value="C")
    else:
        ws.cell(row=row, column=10, value="D")
    add_datas=0


for row in range(2,22):
    if ws.cell(row=row, column=2).value < 5:
        ws.cell(row=row, column=10, value="F")

'''
for row in range(2,22):
    if ws.cell(row=row, column=9).value >= 90:
        ws.cell(row=row, column=10, value="A")

    elif ws.cell(row=row, column=9).value >= 80:
            ws.cell(row=row, column=10, value="B")

    elif ws.cell(row=row, column=9).value >= 70:
            ws.cell(row=row, column=10, value="C")
            
    else:
        ws.cell(row=row, column=10, value="D")


for row in range(2,22):
    if ws.cell(row=row, column=2).value < 5:
        ws.cell(row=row, column=10, value="F")
        
'''
wb.save("scores1.xlsx")
