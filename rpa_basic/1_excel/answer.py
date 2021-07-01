'''
quiz.py의 답입니다. quiz.py에서는 그냥 랜덤으로 제가 만들어봤는데...어렵네요
답은 성적데이터가 주어집니다. 해당 성적 데이터에 맞게 작성하시면 됩니다.

학생 수 : 10
    (1,10,8,10,14,26,12),
    (2,7,3,10,15,24,18),
    (3,9,7,10,8,12,4),
    (4,7,5,10,17,21,18),
    (5,7,8,10,16,25,15),
    (6,3,8,10,8,17,0),
    (7,4,5,10,16,27,18),
    (8,6,9,10,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)
'''

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

#현재까지 작성했던 데이터 넣기
ws.append(["학번", "출석 (10)", "퀴즈1 (10)", "퀴즈2 (10)", "중간고사 (20)", "기말고사 (30)", "프로젝트 (20)"])

scores = [
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)
]

#기존 성적 데이터 넣기
for s in scores:
     ws.append(s)
     
for alp in ["B", "C", "D", "E", "F", "G"]:
    ws.column_dimensions[alp].width = 15


#1. 퀴즈2 점수를 10으로 수정
for idx, cell in enumerate(ws["D"]):
    if idx == 0:
        continue
    cell.value = 10

#등급 매기기
ws["H1"] = "총점"
ws["I1"] = "등급"

for idx, score in enumerate(scores, start=2): #index를 2부터 시작시키기
    sum_val = sum(score[1:]) - score[3] + 10 #총점1
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx,idx) #총점2

    grade = None
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"

    #출석이 5점 미만인 학생은 그냥 F
    if score[1] <5:
        grade = "F"

    ws.cell(row=idx, column=9).value = grade

    
wb.save("scores.xlsx")
    
 
