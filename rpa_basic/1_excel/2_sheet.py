from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()  # 새로운 sheet 기본 이름으로 생성
ws.title = "MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" #rgb 형식으로 값을 넣어주면 됨.


ws1 = wb.create_sheet("Yoursheet")# 주어진 이름으로 sheet 생성
#Sheet, MySheet, YourSheet 순서로 있음.

#sheet를 사이에 생성하고 싶을때
ws2 = wb.create_sheet("NewSheet",2) #2번째에 새 시트가 생성됨.

new_ws = wb["NewSheet"] #dict 형태로 sheet에 접근
print(wb.sheetnames) #모든 sheet 이름 확인

#sheet 복사
new_ws["A1"] = "Test" #A1셀에 해당값 넣기
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"



wb.save("sample.xlsx")


